from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR


class ExcelReport:

    def __init__(self):

        self.workbook = Workbook()

        # Remove default sheet
        self.workbook.remove(self.workbook.active)

    def add_account_sheet(self, account_name, reels):
        reels = sorted(
            reels,
            key=lambda x: x.views,
            reverse=True
        )

        sheet = self.workbook.create_sheet(
            title=account_name[:31]
        )

        headers = [

            "Posted Date",
            "Views",
            "Likes",
            "Views (Text)",
            "Likes (Text)",
            "Reel URL"
        ]

        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for reel in reels:

            sheet.append([
                reel.posted_date.strftime("%Y-%m-%d %H:%M")
                if reel.posted_date else "",
                reel.views,
                reel.likes if reel.likes is not None else "",
                reel.views_text,
                reel.likes_text if reel.likes_text else "",
                reel.url,
            ])

            views_cell = sheet.cell(
                row=sheet.max_row,
                column=2
            )

            views_cell.number_format = "#,##0"
            likes_cell = sheet.cell(
                row=sheet.max_row,
                column=3
            )

            likes_cell.number_format = "#,##0"

            link_cell = sheet.cell(
                row=sheet.max_row,
                column=6
            )

            link_cell.hyperlink = reel.url
            link_cell.style = "Hyperlink"
        # Auto size columns
        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except Exception:

                    pass

            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 4, 80)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    def save(self):

        filename = (
            REPORTS_DIR
            / f"Instagram_Analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        self.workbook.save(filename)

        print()

        print("=" * 80)

        print("Report Saved")

        print(filename)

        print("=" * 80)

        return filename